# Updated for pension/property arrays
"""
Batch retirement planning API endpoint.
Handles bulk scenario calculations with payment verification.
"""
import logging
import io
import csv
from typing import List
from fastapi import APIRouter, HTTPException, Query, status
from fastapi.responses import StreamingResponse

from backend.config import settings
from backend.models.batch_retirement_plan import (
    BatchRetirementPlanInput,
    BatchRetirementPlanOutput,
    ScenarioResult
)
from backend.services.scenario_generator import (
    ScenarioGenerator,
    validate_batch_feasibility
)
from backend.services.retirement_calculator import RetirementCalculator
from backend.services.payment_verifier import SolanaPaymentVerifier

logger = logging.getLogger(__name__)

router = APIRouter()

# Pricing configuration
SOL_PER_RUN = 0.00002  # 0.00002 SOL per scenario (~$0.002-0.004 USD)


@router.post("/calculate-batch-estimate")
async def estimate_batch_calculation(
    batch_input: BatchRetirementPlanInput
) -> dict:
    """
    Estimate cost and time for batch calculation (no payment required).
    
    Returns scenario count, estimated time, and cost in SOL.
    """
    # Validate feasibility
    validation = validate_batch_feasibility(batch_input)
    
    if not validation['feasible']:
        raise HTTPException(
            status_code=400,
            detail=validation['error']
        )
    
    # Calculate cost
    scenario_count = validation['scenario_count']
    cost_sol = scenario_count * SOL_PER_RUN
    
    return {
        "scenario_count": scenario_count,
        "enabled_fields": validation['enabled_fields'],
        "estimated_time_seconds": validation['estimated_time_seconds'],
        "cost_sol": cost_sol,
        "cost_usd_estimate": cost_sol * 150,  # Rough estimate at $150/SOL
        "feasible": True
    }


@router.post("/calculate-batch")
async def calculate_batch_scenarios(
    batch_input: BatchRetirementPlanInput,
    payment_signature: str = Query(..., description="Solana transaction signature"),
    wallet_address: str = Query(..., description="User's wallet address"),
):
    """
    Calculate retirement scenarios with payment verification.
    Returns CSV file with all scenario results.
    """
    logger.info(f"Batch calculation request from wallet {wallet_address[:8]}...")
    
    # 1. Validate feasibility
    validation = validate_batch_feasibility(batch_input)
    if not validation['feasible']:
        raise HTTPException(
            status_code=400,
            detail=f"Batch not feasible: {validation['error']}"
        )
    
    scenario_count = validation['scenario_count']
    logger.info(f"Processing {scenario_count} scenarios ({validation['enabled_fields']} enabled fields)")
    
    # 2. Verify payment
    expected_payment_sol = scenario_count * SOL_PER_RUN
    
    verifier = SolanaPaymentVerifier(
        rpc_url=settings.solana_rpc_url,
        expected_recipient=settings.treasury_wallet,
        expected_amount_sol=expected_payment_sol,
    )
    
    verification = await verifier.verify_transaction(payment_signature, wallet_address)
    
    if not verification.verified:
        raise HTTPException(
            status_code=402,  # Payment Required
            detail=f"Payment verification failed: {verification.error}"
        )
    
    logger.info(f"✅ Payment verified: {verification.amount_sol} SOL (tx: {payment_signature[:8]}...)")
    
    # 3. Generate scenarios
    generator = ScenarioGenerator(batch_input)
    scenarios = generator.generate_scenarios()
    
    logger.info(f"Generated {len(scenarios)} scenario combinations")
    
    # 4. Calculate all scenarios
    calculator = RetirementCalculator()
    results = []
    
    for i, scenario_input in enumerate(scenarios, start=1):
        if i % 100 == 0:
            logger.info(f"Processing scenario {i}/{len(scenarios)}...")
        
        try:
            result = calculator.calculate_plan(scenario_input)
            results.append({
                'scenario_id': i,
                'input': scenario_input,
                'output': result
            })
        except Exception as e:
            logger.error(f"Error calculating scenario {i}: {e}")
            # Continue with other scenarios
            results.append({
                'scenario_id': i,
                'input': scenario_input,
                'error': str(e)
            })
    
    logger.info(f"✅ Completed {len(results)} calculations")
    
    # 5. Generate professional XLSX with 3 tabs
    xlsx_buffer = create_batch_analysis_xlsx(results, batch_input, payment_signature)
    
    # 6. Return as downloadable file
    return StreamingResponse(
        xlsx_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=retirement_analysis_{payment_signature[:8]}.xlsx"
        }
    )


def format_results_as_csv(results: List[dict], batch_input: BatchRetirementPlanInput) -> str:
    """
    Format scenario results as CSV with year-by-year data.
    
    CSV structure:
    scenario_id, retirement_age, rrsp_balance, ..., year, age, rrsp, tfsa, nonreg, total, taxes, warnings
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row
    header = [
        'scenario_id',
        # Input parameters
        'retirement_age', 'rrsp_balance', 'tfsa_balance', 'nonreg_balance',
        'annual_spending', 'monthly_savings',
        'rrsp_real_return', 'tfsa_real_return', 'nonreg_real_return',
        'num_properties', 'num_pensions', 'num_additional_income',
        'cpp_start_age', 'oas_start_age',
        # Year-by-year data
        'year', 'age',
        'rrsp_rrif_balance', 'tfsa_balance_eoy', 'nonreg_balance_eoy',
        'total_balance', 'gross_income', 'taxes_estimated',
        'success', 'final_balance', 'warnings'
    ]

    # Add dynamic columns for pensions
    max_pensions = max(
        (len(s.get('input', {}).get('pensions', [])) if isinstance(s.get('input'), dict) 
         else len(s['input'].pensions) if hasattr(s['input'], 'pensions') and s['input'].pensions 
         else 0)
        for s in results
    ) if results else 0
    
    for i in range(max_pensions):
        header.extend([
            f'pension_{i+1}_monthly',
            f'pension_{i+1}_start_year',
            f'pension_{i+1}_indexing',
            f'pension_{i+1}_end_year'
        ])
    
    # Add dynamic columns for additional income
    max_additional_income = max(
        (len(s.get('input', {}).get('additional_income', [])) if isinstance(s.get('input'), dict) 
         else len(s['input'].additional_income) if hasattr(s['input'], 'additional_income') and s['input'].additional_income 
         else 0)
        for s in results
    ) if results else 0
    
    for i in range(max_additional_income):
        header.extend([
            f'additional_income_{i+1}_monthly',
            f'additional_income_{i+1}_start_year',
            f'additional_income_{i+1}_indexing',
            f'additional_income_{i+1}_end_year'
        ])
    
    writer.writerow(header)
    
    # Data rows (one per scenario per year)
    for result in results:
        scenario_id = result['scenario_id']
        scenario_input = result['input']
        
        if 'error' in result:
            # Write error row
            row = [scenario_id] + ['ERROR'] * (len(header) - 1)
            writer.writerow(row)
            continue
        
        scenario_output = result['output']
        
        # Extract input parameters
        input_params = [
            scenario_input.retirement_age,
            scenario_input.rrsp_balance,
            scenario_input.tfsa_balance,
            scenario_input.non_registered,
            scenario_input.desired_annual_spending,
            scenario_input.monthly_contribution,
            scenario_input.rrsp_real_return,
            scenario_input.tfsa_real_return,
            scenario_input.non_reg_real_return,
            (len(scenario_input.real_estate_holdings) if scenario_input.real_estate_holdings else 0),  # Number of properties
            (len(scenario_input.pensions) if scenario_input.pensions else 0),  # Number of pensions
            (len(scenario_input.additional_income) if scenario_input.additional_income else 0),  # Number of additional income streams
            scenario_input.cpp_start_age,
            scenario_input.oas_start_age,
        ]
        
        # Add pension details
        pensions = scenario_input.pensions if scenario_input.pensions else []
        for i in range(max_pensions):
            if i < len(pensions):
                pension = pensions[i]
                input_params.extend([
                    pension.monthly_amount,
                    pension.start_year,
                    pension.indexing_rate,
                    pension.end_year if pension.end_year else ''
                ])
            else:
                input_params.extend(['', '', '', ''])  # Empty placeholders
        
        # Add additional income details
        additional_incomes = scenario_input.additional_income if scenario_input.additional_income else []
        for i in range(max_additional_income):
            if i < len(additional_incomes):
                income = additional_incomes[i]
                input_params.extend([
                    income.monthly_amount,
                    income.start_year,
                    income.indexing_rate,
                    income.end_year if income.end_year else ''
                ])
            else:
                input_params.extend(['', '', '', ''])  # Empty placeholders
        
        # Summary data
        final_balance = scenario_output.final_balance
        success = scenario_output.success
        warnings_str = "; ".join(scenario_output.warnings) if scenario_output.warnings else ""
        
        # Write one row per year (using correct field names)
        for projection in scenario_output.projections:
            row = [scenario_id] + input_params + [
                projection.year,
                projection.age,
                projection.rrsp_rrif_balance,
                projection.tfsa_balance,
                projection.non_registered_balance,
                projection.total_balance,
                projection.gross_income,
                projection.taxes_estimated,
                success,
                final_balance,
                warnings_str
            ]
            writer.writerow(row)
    
    return output.getvalue()






def create_batch_analysis_xlsx(results: List[dict], batch_input: BatchRetirementPlanInput, payment_signature: str) -> io.BytesIO:
    """
    Generate professional XLSX with 3 tabs:
    1. Summary - Customer-friendly scenario comparison
    2. Detailed Data - Year-by-year breakdown (for B2B clients)
    3. Analysis Guide - Recommendations and formulas
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # ====================================
    # TAB 1: SUMMARY
    # ====================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Headers
    headers = [
        'Scenario', 'Retirement Age', 'Starting RRSP', 'Starting TFSA', 
        'Starting Non-Reg', 'Annual Spending', 'Monthly Savings',
        'RRSP Return %', 'TFSA Return %', 'Non-Reg Return %',
        'CPP Start Age', 'OAS Start Age', '# Pensions', '# Income Streams',
        'Success?', 'Money Lasts To Age', 'Final Balance', 'Total Years'
    ]
    
    # Detect which columns vary across scenarios
    varying_columns = set()
    input_column_indices = list(range(2, 15))  # Columns B-N (inputs only, not outputs)
    
    if len(results) > 1:
        # Check each input column to see if values vary
        for col_idx in input_column_indices:
            values = []
            for result in results:
                if 'error' not in result:
                    scenario_input = result['input']
                    # Map column index to input field
                    if col_idx == 2: values.append(scenario_input.retirement_age)
                    elif col_idx == 3: values.append(scenario_input.rrsp_balance)
                    elif col_idx == 4: values.append(scenario_input.tfsa_balance)
                    elif col_idx == 5: values.append(scenario_input.non_registered)
                    elif col_idx == 6: values.append(scenario_input.desired_annual_spending)
                    elif col_idx == 7: values.append(scenario_input.monthly_contribution)
                    elif col_idx == 8: values.append(scenario_input.rrsp_real_return)
                    elif col_idx == 9: values.append(scenario_input.tfsa_real_return)
                    elif col_idx == 10: values.append(scenario_input.non_reg_real_return)
                    elif col_idx == 11: values.append(scenario_input.cpp_start_age)
                    elif col_idx == 12: values.append(scenario_input.oas_start_age)
                    elif col_idx == 13: values.append(len(scenario_input.pensions) if scenario_input.pensions else 0)
                    elif col_idx == 14: values.append(len(scenario_input.additional_income) if scenario_input.additional_income else 0)
            
            # If values differ, mark as varying
            if len(set(values)) > 1:
                varying_columns.add(col_idx)
    
    # Style headers: black for varying, blue for fixed
    varying_header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fixed_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col_num)
        cell.value = header
        
        # Black header if column varies, blue if fixed
        if col_num in varying_columns:
            cell.fill = varying_header_fill
        else:
            cell.fill = fixed_header_fill
        
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    failure_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    input_column_indices = list(range(2, 12))  # Columns B-K (inputs only)
    
    for row_num, result in enumerate(results, 2):
        scenario_id = result['scenario_id']
        
        if 'error' in result:
            ws_summary.cell(row=row_num, column=1).value = scenario_id
            ws_summary.cell(row=row_num, column=2).value = 'ERROR'
            continue
        
        scenario_input = result['input']
        scenario_output = result['output']
        
        # Calculate money_lasts_to_age
        projections = scenario_output.projections
        money_lasts_to_age = scenario_input.life_expectancy
        for proj in projections:
            if proj.total_balance <= 0:
                money_lasts_to_age = proj.age
                break
        
        total_years = money_lasts_to_age - scenario_input.retirement_age
        success = scenario_output.success
        
        # Write row data
        row_data = [
            scenario_id,
            scenario_input.retirement_age,
            scenario_input.rrsp_balance,
            scenario_input.tfsa_balance,
            scenario_input.non_registered,
            scenario_input.desired_annual_spending,
            scenario_input.monthly_contribution,
            f"{scenario_input.rrsp_real_return * 100:.1f}%",
            f"{scenario_input.tfsa_real_return * 100:.1f}%",
            f"{scenario_input.non_reg_real_return * 100:.1f}%",
            scenario_input.cpp_start_age,
            scenario_input.oas_start_age,
            len(scenario_input.pensions) if scenario_input.pensions else 0,
            len(scenario_input.additional_income) if scenario_input.additional_income else 0,
            'YES' if success else 'NO',
            money_lasts_to_age,
            scenario_output.final_balance,
            total_years
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_num, column=col_num)
            cell.value = value
            
            # Center align return % columns and Success column
            if col_num in [8, 9, 10, 15]:  # RRSP Return %, TFSA Return %, Non-Reg Return %, Success
                cell.alignment = Alignment(horizontal='center')
            
            # Currency formatting for money columns
            if col_num in [3, 4, 5, 6, 7, 17]:  # RRSP, TFSA, Non-Reg, Spending, Savings, Final Balance
                cell.number_format = '"$"#,##0'
            
            # Color code success/failure
            if col_num == 15:  # Success column
                if success:
                    cell.fill = success_fill
                    cell.font = Font(bold=True, color="006100")
                else:
                    cell.fill = failure_fill
                    cell.font = Font(bold=True, color="9C0006")
        
        # Highlight cells that changed from previous row (orange)
        change_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        
        if row_num > 2:  # Compare to previous row
            prev_result = results[row_num - 3]  # -3 because row_num starts at 2
            if 'error' not in prev_result and 'error' not in result:
                prev_input = prev_result['input']
                curr_input = scenario_input
                
                # Check each input column for changes
                for col_num, value in enumerate(row_data, 1):
                    if col_num in input_column_indices:  # Only highlight input columns
                        changed = False
                        
                        if col_num == 2: changed = (prev_input.retirement_age != curr_input.retirement_age)
                        elif col_num == 3: changed = (prev_input.rrsp_balance != curr_input.rrsp_balance)
                        elif col_num == 4: changed = (prev_input.tfsa_balance != curr_input.tfsa_balance)
                        elif col_num == 5: changed = (prev_input.non_registered != curr_input.non_registered)
                        elif col_num == 6: changed = (prev_input.desired_annual_spending != curr_input.desired_annual_spending)
                        elif col_num == 7: changed = (prev_input.monthly_contribution != curr_input.monthly_contribution)
                        elif col_num == 8: changed = (prev_input.cpp_start_age != curr_input.cpp_start_age)
                        elif col_num == 9: changed = (prev_input.oas_start_age != curr_input.oas_start_age)
                        elif col_num == 13: 
                            prev_pensions = len(prev_input.pensions) if prev_input.pensions else 0
                            curr_pensions = len(curr_input.pensions) if curr_input.pensions else 0
                            changed = (prev_pensions != curr_pensions)
                        elif col_num == 14:
                            prev_income = len(prev_input.additional_income) if prev_input.additional_income else 0
                            curr_income = len(curr_input.additional_income) if curr_input.additional_income else 0
                            changed = (prev_income != curr_income)
                        
                        if changed:
                            cell = ws_summary.cell(row=row_num, column=col_num)
                            cell.fill = change_fill
        
        # Add alternating row shading for better readability (but don't override orange highlights)
        if row_num % 2 == 0:  # Even rows
            alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws_summary.cell(row=row_num, column=col)
                # Don't override success/failure colors or change highlights
                if col != 15 and not (cell.fill and cell.fill.start_color.rgb in ['00FFC7CE', '00C6EFCE', '00FFE0B2']):
                    cell.fill = alt_fill
    
    # Optimized column widths for readability
    column_widths = {
        'A': 10,  # Scenario
        'B': 14,  # Retirement Age
        'C': 15,  # Starting RRSP
        'D': 15,  # Starting TFSA
        'E': 17,  # Starting Non-Reg
        'F': 16,  # Annual Spending
        'G': 15,  # Monthly Savings
        'H': 14,  # RRSP Return %
        'I': 13,  # TFSA Return %
        'J': 16,  # Non-Reg Return %
        'K': 14,  # CPP Start Age
        'L': 15,  # OAS Start Age (wider for header)
        'M': 11,  # # Pensions
        'N': 16,  # # Income Streams
        'O': 10,  # Success?
        'P': 18,  # Money Lasts To Age
        'Q': 15,  # Final Balance
        'R': 12,  # Total Years
    }
    
    for col_letter, width in column_widths.items():
        ws_summary.column_dimensions[col_letter].width = width
    
    # Freeze top row for easier scrolling
    ws_summary.freeze_panes = 'A2'
    
    # ====================================
    # TAB 2: DETAILED DATA
    # ====================================
    ws_detailed = wb.create_sheet("Detailed Data")
    
    # Define header styling for detailed tab (reuse from summary)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Use existing CSV format function to get detailed data
    detailed_csv = format_results_as_csv(results, batch_input)
    
    # Parse CSV and write to sheet
    import csv
    csv_reader = csv.reader(io.StringIO(detailed_csv))
    for row_num, row_data in enumerate(csv_reader, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_detailed.cell(row=row_num, column=col_num)
            try:
                cell.value = float(value)
            except (ValueError, TypeError):
                cell.value = value
            
            # Header styling
            if row_num == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    # ====================================
    # TAB 3: ANALYSIS GUIDE
    # ====================================
    ws_guide = wb.create_sheet("Analysis Guide")
    ws_guide.column_dimensions['A'].width = 100  # Wider for better readability
    
    guide_content = [
        ("📊 RETIREMENT ANALYSIS RESULTS", Font(bold=True, size=16, color="1F4E78")),
        ("", None),
        ("═" * 60, None),
        ("", None),
        ("🎯 HOW TO FIND YOUR BEST SCENARIO:", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("Your goal: Retire as EARLY as possible while money lasts to life expectancy!", None),
        ("", None),
        ("Look for scenarios marked 'YES' in the Success column (Summary tab).", None),
        ("Among successful scenarios, choose the one with:", None),
        ("   1. ✅ EARLIEST retirement age (retire sooner!)", None),
        ("   2. ✅ HIGHEST annual spending (if tied on age)", None),
        ("   3. ✅ LARGEST final balance (safety cushion)", None),
        ("", None),
        ("═" * 60, None),
        ("", None),
        ("💡 KEY INSIGHTS:", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("✅ SUCCESS = Money lasts to your life expectancy or beyond", None),
        ("❌ FAILURE = Money runs out too early", None),
        ("", None),
        ("═" * 60, None),
        ("", None),
        ("📈 IF NO SCENARIOS SUCCEED, TRY:", Font(bold=True, size=12, color="C00000")),
        ("", None),
        ("   1. Increase monthly savings before retirement", None),
        ("   2. Reduce annual spending in retirement", None),
        ("   3. Delay retirement by 1-2 years", None),
        ("   4. Adjust investment returns (risk tolerance)", None),
        ("", None),
        ("═" * 60, None),
        ("", None),
        ("🔍 TABS EXPLAINED:", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("• Summary - Quick scenario comparison (start here!)", None),
        ("• Detailed Data - Year-by-year breakdown (for detailed analysis)", None),
        ("• Analysis Guide - This tab (how to use the results)", None),
        ("", None),
        ("═" * 60, None),
        ("", None),
        (f"📅 Generated: {payment_signature[:8]}...", Font(italic=True, size=9, color="7F7F7F")),
    ]
    
    # Define border style for separators
    thick_border = Border(bottom=Side(style='thick', color='4472C4'))
    
    for row_num, (text, font) in enumerate(guide_content, 1):
        cell = ws_guide.cell(row=row_num, column=1)
        cell.value = text
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Add bottom border for section separators (replace ═══ lines)
        if text and text.startswith("═"):
            cell.value = ""  # Remove the text separator
            cell.border = thick_border
    
    # Save to BytesIO
    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)
    
    return xlsx_buffer


def format_summary_as_csv(results: List[dict]) -> str:
    """
    Format scenario results as summary CSV - one row per scenario.
    Customer-friendly view with key metrics only.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Simple, readable headers
    header = [
        'Scenario',
        'Retirement Age',
        'Starting RRSP',
        'Starting TFSA',
        'Starting Non-Reg',
        'Annual Spending',
        'Monthly Savings',
        'CPP Start Age',
        'OAS Start Age',
        '# Pensions',
        '# Income Streams',
        'Success?',
        'Money Lasts To Age',
        'Final Balance',
        'Total Years',
        'Warnings'
    ]
    writer.writerow(header)
    
    # One row per scenario
    for result in results:
        scenario_id = result['scenario_id']
        
        if 'error' in result:
            row = [scenario_id, 'ERROR'] + [''] * (len(header) - 2)
            writer.writerow(row)
            continue
        
        scenario_input = result['input']
        scenario_output = result['output']
        
        # Calculate money_lasts_to_age
        projections = scenario_output.projections
        money_lasts_to_age = scenario_input.life_expectancy
        for proj in projections:
            if proj.total_balance <= 0:
                money_lasts_to_age = proj.age
                break
        
        total_years = money_lasts_to_age - scenario_input.retirement_age
        
        row = [
            scenario_id,
            scenario_input.retirement_age,
            f'${scenario_input.rrsp_balance:,.0f}',
            f'${scenario_input.tfsa_balance:,.0f}',
            f'${scenario_input.non_registered:,.0f}',
            f'${scenario_input.desired_annual_spending:,.0f}',
            f'${scenario_input.monthly_contribution:,.0f}',
            scenario_input.cpp_start_age,
            scenario_input.oas_start_age,
            len(scenario_input.pensions) if scenario_input.pensions else 0,
            len(scenario_input.additional_income) if scenario_input.additional_income else 0,
            'YES' if scenario_output.success else 'NO',
            money_lasts_to_age,
            f'${scenario_output.final_balance:,.0f}',
            total_years,
            '; '.join(scenario_output.warnings) if scenario_output.warnings else ''
        ]
        writer.writerow(row)
    
    return output.getvalue()


@router.post("/calculate-batch-test")
async def calculate_batch_test(
    batch_input: BatchRetirementPlanInput
):
    """
    TEST ONLY: Calculate batch without payment verification.
    Remove this endpoint before production deployment!
    """
    logger.warning("⚠️ Using TEST endpoint - no payment verification!")
    
    # Validate feasibility
    validation = validate_batch_feasibility(batch_input)
    if not validation['feasible']:
        raise HTTPException(status_code=400, detail=validation['error'])
    
    # Generate scenarios
    generator = ScenarioGenerator(batch_input)
    scenarios = generator.generate_scenarios()
    
    logger.info(f"Calculating {len(scenarios)} scenarios...")
    
    # Calculate all scenarios
    calculator = RetirementCalculator()
    results = []
    
    for i, scenario_input in enumerate(scenarios, start=1):
        try:
            result = calculator.calculate_plan(scenario_input)
            results.append({
                'scenario_id': i,
                'input': scenario_input,
                'output': result
            })
        except Exception as e:
            logger.error(f"Error in scenario {i}: {e}")
            results.append({
                'scenario_id': i,
                'input': scenario_input,
                'error': str(e)
            })
    
    # Format as CSV
    csv_content = format_results_as_csv(results, batch_input)
    
    # Return as file
    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=test_scenarios.xlsx"}
    )
